"""Shared file handler for Slack and Telegram bots.

Downloads user-submitted files, stores them in a structured workspace,
and performs best-effort text extraction for context injection into
agent prompts.

Not a module — a utility used by both bot modules.
"""

from __future__ import annotations

import logging
import mimetypes
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_logger = logging.getLogger("arcagent.file_handler")

# Safe filename pattern — strip everything except alphanumerics, hyphens, dots, underscores
_UNSAFE_CHARS = re.compile(r"[^\w\-.]")

# Default maximum file size in bytes (20 MB)
_DEFAULT_MAX_FILE_SIZE = 20 * 1024 * 1024

# Text-like MIME type prefixes and extensions
_TEXT_EXTENSIONS = frozenset({
    ".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".xml",
    ".py", ".js", ".ts", ".html", ".css", ".sh", ".bash",
    ".toml", ".ini", ".cfg", ".conf", ".log", ".sql",
    ".rs", ".go", ".java", ".c", ".cpp", ".h", ".hpp",
    ".rb", ".php", ".swift", ".kt", ".r", ".m",
})

_MAX_TEXT_EXTRACT_BYTES = 50_000  # Cap extracted text at ~50KB


def _sanitize_filename(name: str) -> str:
    """Sanitize a filename for safe local storage."""
    # Strip path components
    name = Path(name).name
    # Replace unsafe chars with underscores
    name = _UNSAFE_CHARS.sub("_", name)
    # Collapse multiple underscores
    name = re.sub(r"_+", "_", name)
    # Prepend date prefix
    date_prefix = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    return f"{date_prefix}_{name}"


class FileHandler:
    """Download, store, and extract text from user-submitted files.

    Args:
        workspace: Agent workspace directory (contains files/ subdirectory).
        max_file_size: Maximum allowed file size in bytes.
    """

    def __init__(
        self,
        workspace: Path,
        max_file_size: int = _DEFAULT_MAX_FILE_SIZE,
    ) -> None:
        self._workspace = workspace
        self._max_file_size = max_file_size
        self._inbox_dir = workspace / "files" / "inbox"
        self._inbox_dir.mkdir(parents=True, exist_ok=True)

    @property
    def inbox_dir(self) -> Path:
        return self._inbox_dir

    async def download_from_bytes(
        self,
        data: bytes,
        filename: str,
    ) -> Path | None:
        """Store raw bytes as a file in the inbox.

        Used when the platform API gives us the bytes directly
        (e.g., python-telegram-bot's download_as_bytearray).
        """
        if len(data) > self._max_file_size:
            _logger.warning(
                "File %s exceeds max size (%d > %d); skipping",
                filename,
                len(data),
                self._max_file_size,
            )
            return None

        safe_name = _sanitize_filename(filename)
        dest = self._inbox_dir / safe_name

        # Avoid overwriting — append suffix if needed
        counter = 1
        while dest.exists():
            stem = dest.stem
            dest = self._inbox_dir / f"{stem}_{counter}{dest.suffix}"
            counter += 1

        dest.write_bytes(data)
        _logger.info("Stored file: %s (%d bytes)", dest.name, len(data))
        return dest

    async def download_from_url(
        self,
        url: str,
        headers: dict[str, str] | None = None,
        filename: str | None = None,
    ) -> Path | None:
        """Download a file from a URL into the inbox directory.

        Used for Slack file downloads (private URL + bot token header).
        """
        try:
            import httpx
        except ImportError:
            _logger.warning("httpx not installed; cannot download file from URL")
            return None

        try:
            async with httpx.AsyncClient(follow_redirects=True) as client:
                resp = await client.get(url, headers=headers or {}, timeout=60.0)
                resp.raise_for_status()

                # Check size
                data = resp.content
                if len(data) > self._max_file_size:
                    _logger.warning(
                        "Downloaded file exceeds max size (%d > %d); discarding",
                        len(data),
                        self._max_file_size,
                    )
                    return None

                # Determine filename
                if not filename:
                    filename = url.rsplit("/", 1)[-1].split("?")[0] or "download"

                return await self.download_from_bytes(data, filename)

        except Exception:
            _logger.exception("Failed to download file from URL: %s", url[:100])
            return None

    def extract_text(self, path: Path) -> str | None:
        """Best-effort text extraction from a file.

        Returns extracted text content, or None if extraction
        isn't possible or the file type isn't supported.
        """
        suffix = path.suffix.lower()
        mime_type, _ = mimetypes.guess_type(str(path))

        # Plain text / code / structured data
        if suffix in _TEXT_EXTENSIONS or (mime_type and mime_type.startswith("text/")):
            return self._extract_text_file(path)

        # PDF
        if suffix == ".pdf":
            return self._extract_pdf(path)

        # Word documents
        if suffix in (".docx", ".doc"):
            return self._extract_docx(path)

        # Excel
        if suffix in (".xlsx", ".xls"):
            return self._extract_xlsx(path)

        # Images — no extraction, just metadata
        if suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp"):
            size_kb = path.stat().st_size / 1024
            return f"[Image: {path.name}, {size_kb:.0f}KB]"

        # Unknown type — just metadata
        size_kb = path.stat().st_size / 1024
        return f"[File: {path.name}, {size_kb:.0f}KB, type: {mime_type or 'unknown'}]"

    def build_context(self, path: Path, extracted_text: str | None) -> str:
        """Format file info + extracted text as a prompt context block."""
        size_kb = path.stat().st_size / 1024
        lines = [
            f"---",
            f"**Attached file:** `{path.name}` ({size_kb:.1f} KB)",
            f"**Stored at:** `{path}`",
        ]

        if extracted_text:
            # Truncate very long extractions
            if len(extracted_text) > _MAX_TEXT_EXTRACT_BYTES:
                extracted_text = extracted_text[:_MAX_TEXT_EXTRACT_BYTES] + "\n... [truncated]"
            lines.append(f"**Content:**\n```\n{extracted_text}\n```")
        else:
            lines.append("*(Binary file — content not extracted)*")

        lines.append("---")
        return "\n".join(lines)

    def store_team(self, path: Path, agent_name: str, team_root: Path) -> Path:
        """Copy a file from inbox to the team shared directory.

        Returns the destination path.
        """
        dest_dir = team_root / "shared" / "files" / agent_name
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / path.name

        import shutil
        shutil.copy2(path, dest)
        _logger.info("Stored team file: %s → %s", path.name, dest)
        return dest

    # ── Private extraction helpers ────────────────────────────────

    def _extract_text_file(self, path: Path) -> str | None:
        """Read a plain text file."""
        try:
            text = path.read_text(errors="replace")
            if len(text) > _MAX_TEXT_EXTRACT_BYTES:
                text = text[:_MAX_TEXT_EXTRACT_BYTES] + "\n... [truncated]"
            return text
        except Exception:
            _logger.debug("Failed to read text file: %s", path)
            return None

    def _extract_pdf(self, path: Path) -> str | None:
        """Extract text from PDF using pdfplumber or PyPDF2."""
        # Try pdfplumber first (better extraction quality)
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                pages = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        pages.append(text)
                return "\n\n".join(pages) if pages else None
        except ImportError:
            pass
        except Exception:
            _logger.debug("pdfplumber failed for %s", path)

        # Fallback to PyPDF2
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(path))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages) if pages else None
        except ImportError:
            _logger.debug("No PDF library available (pdfplumber or PyPDF2)")
            return None
        except Exception:
            _logger.debug("PyPDF2 failed for %s", path)
            return None

    def _extract_docx(self, path: Path) -> str | None:
        """Extract text from Word documents."""
        try:
            from docx import Document
            doc = Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else None
        except ImportError:
            _logger.debug("python-docx not installed; cannot extract .docx")
            return None
        except Exception:
            _logger.debug("Failed to extract docx: %s", path)
            return None

    def _extract_xlsx(self, path: Path) -> str | None:
        """Extract text from Excel spreadsheets."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            sheets = []
            for ws in wb.worksheets:
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append("\t".join(cells))
                if rows:
                    sheets.append(f"[Sheet: {ws.title}]\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(sheets) if sheets else None
        except ImportError:
            _logger.debug("openpyxl not installed; cannot extract .xlsx")
            return None
        except Exception:
            _logger.debug("Failed to extract xlsx: %s", path)
            return None
